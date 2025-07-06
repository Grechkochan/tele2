import warnings
warnings.filterwarnings("ignore", "Unknown extension.*", UserWarning)
import psycopg2
import re
from datetime import datetime
from openpyxl import load_workbook
from config import DB_HOST, DB_PORT, DB_NAME, DB_USER, DB_PASSWORD

TEMPLATE_PATH = "отчет.xlsx"
OUTPUT_PATH   = "OPEX_filled.xlsx"
SHEET_NAME    = "Opex"
START_ROW     = 11
SUMMARY_ROW   = 14
SUMMARY_COL   = 38

def fetch_tasks():
    conn = psycopg2.connect(
        host=DB_HOST, port=DB_PORT,
        dbname=DB_NAME, user=DB_USER, password=DB_PASSWORD
    )
    cur = conn.cursor()
    cur.execute("""
        SELECT
            T.basestation,
            T.tasknum,
            T.status,
            COALESCE(W.workerFIO, '') AS workerFIO,
            T.datetime,
            T.datetimeacc,
            T.datetimeclose,
            T.responsible_person,
            T.close_code,
            T.work_type,
            T.quantity
        FROM Tasks AS T
        LEFT JOIN Workers AS W
          ON T.worker = W.tgId
        ORDER BY T.datetime
    """)
    rows = cur.fetchall()
    conn.close()
    return rows

def fetch_tasks_for_worker(worker_id, start_date, end_date):
    conn = psycopg2.connect(
        host=DB_HOST, port=DB_PORT,
        dbname=DB_NAME, user=DB_USER, password=DB_PASSWORD
    )
    cur = conn.cursor()
    cur.execute("""
        SELECT
            T.basestation,
            T.tasknum,
            T.status,
            COALESCE(W.workerFIO, '') AS workerFIO,
            T.datetime,
            T.datetimeacc,
            T.datetimeclose,
            T.responsible_person,
            T.close_code,
            T.work_type,
            T.quantity
        FROM Tasks AS T
        LEFT JOIN Workers AS W
          ON T.worker = W.tgId
        WHERE T.worker = %s
          AND DATE(T.datetimeclose) BETWEEN %s AND %s
        ORDER BY T.datetime
    """, (str(worker_id), start_date, end_date))
    rows = cur.fetchall()
    conn.close()
    return rows

def fetch_tasks_for_all(start_date, end_date):
    conn = psycopg2.connect(
        host=DB_HOST, port=DB_PORT,
        dbname=DB_NAME, user=DB_USER, password=DB_PASSWORD
    )
    cur = conn.cursor()
    cur.execute("""
        SELECT
            T.basestation,
            T.tasknum,
            T.status,
            COALESCE(W.workerFIO, '') AS workerFIO,
            T.datetime,
            T.datetimeacc,
            T.datetimeclose,
            T.responsible_person,
            T.close_code,
            T.work_type,
            T.quantity
        FROM Tasks AS T
        LEFT JOIN Workers AS W
          ON T.worker = W.tgId
        WHERE DATE(T.datetimeclose) BETWEEN %s AND %s
        ORDER BY T.datetime
    """, (start_date, end_date))
    rows = cur.fetchall()
    conn.close()
    return rows

def collect_formulas(ws, row):
    formulas = {}
    for cell in ws[row]:
        if isinstance(cell.value, str) and cell.value.startswith("="):
            formulas[cell.column] = cell.value
    return formulas

def shift_formula(formula: str, offset: int) -> str:
    def repl(m):
        col = m.group("col")
        row = int(m.group("row"))
        if col.endswith("$"):
            return f"{col}{row}"

        return f"{col}{row + offset}"

    pattern = re.compile(r"(?P<col>\$?[A-Z]+\$?)(?P<row>\d+)")
    return pattern.sub(repl, formula)

def parse_dt(val, default=None):
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(val, fmt)
            except:
                pass
        try:
            return datetime.fromisoformat(val)
        except:
            return default
    return default

def expand_task(rec):
    bs, num, status, worker, open_s, acc_s, close_s, resp, close_codes, wtype, quantities = rec
    status_l = (status or "").strip().lower()
    closed   = status_l.startswith("закр")
    canceled = "отмен" in status_l 

    if not (closed or canceled):
        return []

    opened    = parse_dt(open_s)
    arrived   = parse_dt(acc_s, default=opened)
    closed_dt = parse_dt(close_s)
    if not opened or not closed_dt:
        return []

    def before_17(dt): return dt.hour < 17
    wtype_l = (wtype or "").strip().lower()
    out = []

    if wtype_l == "ппр":
        if closed:
            codes = close_codes if isinstance(close_codes, list) else [close_codes]
            qtys  = quantities  if isinstance(quantities,  list) else [quantities]
            for idx, code in enumerate(codes):
                qty = qtys[idx] if idx < len(qtys) else None
                out.append((bs, num, opened, arrived, closed_dt,
                            worker, resp, code, qty))
        return out
    
    if wtype_l == "авр":
        if canceled:
            part_code = "3.4.1." if before_17(opened) else "3.4.2."
            out.append((bs, num, opened, arrived, closed_dt,
                        worker, resp, part_code, None))
            return out
        
        code = close_codes[0] if close_codes else ""
        if closed and code.strip().lower() == "генерация":
            total_h = (closed_dt - opened).total_seconds() / 3600
            n72 = int(total_h // 72)
            rem = total_h - n72 * 72
            n24 = int(rem // 24)
            h   = int(rem - n24 * 24)
            for part_code, qty in [("3.3.3.", n72), ("3.3.2.", n24), ("3.3.1.", h)]:
                if qty:
                    out.append((bs, num, opened, arrived, closed_dt,
                                worker, resp, part_code, qty))
        return out
    
    if closed:
        diff_h = (closed_dt - arrived).total_seconds() / 3600
        before_17 = arrived.hour < 17
        if diff_h <= 5:
            part_code = "3.1.1." if before_17 else "3.1.2."
            out.append((bs, num, opened, arrived, closed_dt,
                        worker, resp, part_code, None))
        else:
            c1 = "3.1.1." if before_17 else "3.1.2."
            c2 = "3.2.1." if before_17 else "3.2.2."
            out.append((bs, num, opened, arrived, closed_dt,
                        worker, resp, c1, None))
            out.append((bs, num, opened, arrived, closed_dt,
                        worker, resp, c2, None))
    return out

def fill_report():
    wb = load_workbook(filename=TEMPLATE_PATH)
    ws = wb[SHEET_NAME]
    raw = fetch_tasks()
    tasks = []
    for rec in raw:
        tasks.extend(expand_task(rec))
    if not tasks:
        print("Нет задач для вставки.")
        return
    formulas = collect_formulas(ws, START_ROW)
    n = len(tasks)
    ws.insert_rows(START_ROW + 1, amount=n)
    for idx, (bs, num, opened, arrived, closed_dt, worker, resp, code, qty) in enumerate(tasks, start=1):
        row = START_ROW + idx
        ws.cell(row=row, column=1, value=idx)
        for col, f in formulas.items():
            ws.cell(row=row, column=col, value=shift_formula(f, idx))
        ws.cell(row=row, column=2,  value=bs)
        ws.cell(row=row, column=6,  value=num)
        ws.cell(row=row, column=7,  value=opened)
        ws.cell(row=row, column=8,  value=arrived)
        ws.cell(row=row, column=9,  value=closed_dt)
        ws.cell(row=row, column=10, value=worker)
        ws.cell(row=row, column=11, value=resp)
        ws.cell(row=row, column=12, value=code)
        ws.cell(row=row, column=15, value=qty)
        if code == "3.3.1." and qty is not None:
            ws.cell(row=row, column=17, value=qty)
        elif code == "3.3.2." and qty is not None:
            ws.cell(row=row, column=18, value=qty)
        elif code == "3.3.3." and qty is not None:
            ws.cell(row=row, column=19, value=qty)

    data_start   = START_ROW + 1       
    data_end     = START_ROW + n        
    new_sum_row  = SUMMARY_ROW + n      
    ws.cell(row=new_sum_row, column=SUMMARY_COL).value = (
        f"=SUM(AL{data_start}:AL{data_end})"
    )

    wb.save(OUTPUT_PATH)
    print(f"Готово! Отчёт сохранён в {OUTPUT_PATH}")

def fill_report_for(worker_id: str, start_date: str, end_date: str, 
                    template: str = TEMPLATE_PATH, output: str = OUTPUT_PATH):

    wb = load_workbook(filename=template)
    ws = wb[SHEET_NAME]

    raw = fetch_tasks_for_worker(str(worker_id), start_date, end_date)
    tasks = []
    for rec in raw:
        tasks.extend(expand_task(rec))

    if not tasks:
        print("Нет задач для вставки.")
        return False
    
    formulas = collect_formulas(ws, START_ROW)
    n = len(tasks)
    ws.insert_rows(START_ROW + 1, amount=n)
    for idx, (bs, num, opened, arrived, closed_dt, worker, resp, code, qty) in enumerate(tasks, start=1):
        row = START_ROW + idx
        ws.cell(row=row, column=1, value=idx)
        for col, f in formulas.items():
            ws.cell(row=row, column=col, value=shift_formula(f, idx))
        ws.cell(row=row, column=2,  value=bs)
        ws.cell(row=row, column=6,  value=num)
        ws.cell(row=row, column=7,  value=opened)
        ws.cell(row=row, column=8,  value=arrived)
        ws.cell(row=row, column=9,  value=closed_dt)
        ws.cell(row=row, column=10, value=worker)
        ws.cell(row=row, column=11, value=resp)
        ws.cell(row=row, column=12, value=code)
        ws.cell(row=row, column=15, value=qty)
        if code == "3.3.1." and qty is not None:
            ws.cell(row=row, column=17, value=qty)
        elif code == "3.3.2." and qty is not None:
            ws.cell(row=row, column=18, value=qty)
        elif code == "3.3.3." and qty is not None:
            ws.cell(row=row, column=19, value=qty)
            
    data_start  = START_ROW + 1
    data_end    = START_ROW + n
    sum_cell    = ws.cell(row=SUMMARY_ROW + n, column=SUMMARY_COL)
    sum_cell.value = f"=SUM(AL{data_start}:AL{data_end})"

    wb.save(output)
    print(f"Готово! Отчёт для {worker_id} сохранён в {output}")
    return True

def fill_report_for_all(start_date: str,
                        end_date:   str,
                        template:   str = TEMPLATE_PATH,
                        output:     str = OUTPUT_PATH) -> bool:
    wb = load_workbook(filename=template)
    ws = wb[SHEET_NAME]

    raw = fetch_tasks_for_all(start_date, end_date)
    if not raw:
        return False
    
    tasks = []
    for rec in raw:
        tasks.extend(expand_task(rec))
    formulas = collect_formulas(ws, START_ROW)
    n = len(tasks)
    ws.insert_rows(START_ROW + 1, amount=n)
    for idx, (bs, num, opened, arrived, closed_dt,
              worker, resp, code, qty) in enumerate(tasks, start=1):
        row = START_ROW + idx
        ws.cell(row=row, column=1, value=idx)
        for col, f in formulas.items():
            ws.cell(row=row, column=col, value=shift_formula(f, idx))
        ws.cell(row=row, column=2,  value=bs)
        ws.cell(row=row, column=6,  value=num)
        ws.cell(row=row, column=7,  value=opened)
        ws.cell(row=row, column=8,  value=arrived)
        ws.cell(row=row, column=9,  value=closed_dt)
        ws.cell(row=row, column=10, value=worker)
        ws.cell(row=row, column=11, value=resp)
        ws.cell(row=row, column=12, value=code)
        ws.cell(row=row, column=15, value=qty)
        if code == "3.3.1." and qty is not None:
            ws.cell(row=row, column=17, value=qty)
        elif code == "3.3.2." and qty is not None:
            ws.cell(row=row, column=18, value=qty)
        elif code == "3.3.3." and qty is not None:
            ws.cell(row=row, column=19, value=qty)

    data_start = START_ROW + 1
    data_end   = START_ROW + n
    ws.cell(
        row=SUMMARY_ROW + n,
        column=SUMMARY_COL,
        value=f"=SUM(AL{data_start}:AL{data_end})"
    )
    wb.save(output)
    return True

if __name__ == "__main__":
    fill_report()
